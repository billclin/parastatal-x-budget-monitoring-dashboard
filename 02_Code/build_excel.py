"""
Parastatal X - Excel Workbook Builder
Creates a professionally formatted 23-sheet workbook.
"""
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
import os

from pathlib import Path
ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "01_Data"
OUT  = ROOT / "01_Data"

# ── Colour palette ──────────────────────────────────────────────────────────
FOREST_GREEN   = "1B4332"
MUTED_GOLD     = "B7950B"
WHITE          = "FFFFFF"
LIGHT_GREY     = "F2F2F2"
CHARCOAL       = "2C3E50"
LIGHT_GREEN    = "D5E8D4"
LIGHT_GOLD     = "FFF2CC"
AMBER          = "E67E22"
RED_ACCENT     = "C0392B"

def hdr_fill(color=FOREST_GREEN):
    return PatternFill("solid", fgColor=color)

def hdr_font(color=WHITE, bold=True, size=10):
    return Font(name="Calibri", color=color, bold=bold, size=size)

def body_font(size=9):
    return Font(name="Calibri", size=size, color=CHARCOAL)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def center_align():
    return Alignment(horizontal="center", vertical="center", wrap_text=False)

def wrap_align():
    return Alignment(wrap_text=True, vertical="top")

KSH_FMT  = '#,##0'
PCT_FMT  = '0.0%'
DATE_FMT = 'DD-MMM-YYYY'

wb = Workbook()
wb.remove(wb.active)  # remove default sheet

def add_df_sheet(wb, sheet_name, df, freeze_col=1, freeze_row=2,
                 currency_cols=None, date_cols=None, pct_cols=None,
                 table_name=None):
    """Add a DataFrame as a formatted sheet with table styling."""
    ws = wb.create_sheet(title=sheet_name[:31])
    
    # Header row
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = hdr_fill()
        cell.font = hdr_font()
        cell.alignment = center_align()
        cell.border = thin_border()
        ws.row_dimensions[1].height = 30
    
    # Data rows
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        fill = PatternFill("solid", fgColor=LIGHT_GREY if row_idx % 2 == 0 else WHITE)
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.font = body_font()
            cell.border = thin_border()
            cell.alignment = Alignment(vertical="center")
    
    # Column widths
    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = max(len(str(col_name)),
                      df.iloc[:, col_idx-1].astype(str).str.len().max() if len(df)>0 else 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len * 1.1 + 2, 10), 40)
    
    # Apply number formats
    if currency_cols:
        for col_name in currency_cols:
            if col_name in df.columns:
                col_idx = df.columns.get_loc(col_name) + 1
                for row_idx in range(2, len(df) + 2):
                    ws.cell(row=row_idx, column=col_idx).number_format = KSH_FMT
    if date_cols:
        for col_name in date_cols:
            if col_name in df.columns:
                col_idx = df.columns.get_loc(col_name) + 1
                for row_idx in range(2, len(df) + 2):
                    ws.cell(row=row_idx, column=col_idx).number_format = DATE_FMT
    if pct_cols:
        for col_name in pct_cols:
            if col_name in df.columns:
                col_idx = df.columns.get_loc(col_name) + 1
                for row_idx in range(2, len(df) + 2):
                    ws.cell(row=row_idx, column=col_idx).number_format = PCT_FMT
    
    # Excel table
    if len(df) > 0 and table_name:
        tbl_ref = f"A1:{get_column_letter(len(df.columns))}{len(df)+1}"
        tbl = Table(displayName=table_name[:255], ref=tbl_ref)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True,
                               showColumnStripes=False)
        tbl.tableStyleInfo = style
        ws.add_table(tbl)
    
    # Freeze panes
    ws.freeze_panes = ws.cell(row=freeze_row, column=freeze_col+1)
    return ws

# ── 1. README ───────────────────────────────────────────────────────────────
print("  Building README sheet...")
ws_readme = wb.create_sheet("README")
ws_readme.sheet_view.showGridLines = False
content = [
    ("PARASTATAL X", "DYNAMIC BUDGET MONITORING, FORECASTING AND MANAGEMENT DASHBOARD"),
    ("", ""),
    ("DISCLAIMER", "All data in this workbook is entirely synthetic and generated for portfolio demonstration purposes only."),
    ("", "No real persons, institutions, suppliers, account numbers or government data are represented."),
    ("", ""),
    ("PROJECT TITLE", "Parastatal X Dynamic Budget Monitoring, Forecasting and Management Dashboard"),
    ("SUBTITLE", "A simulated Kenyan public-sector financial analytics solution for near-real-time budget control,"),
    ("", "expenditure monitoring, forecasting and management decision-making"),
    ("", ""),
    ("REPORTING PERIOD", "FY 2023/2024 to FY 2025/2026 (Active Year: FY 2025/2026)"),
    ("CUT-OFF DATE", "31 May 2026"),
    ("CURRENCY", "Kenya Shillings (KSh)"),
    ("FINANCIAL YEAR", "1 July to 30 June"),
    ("", ""),
    ("WORKSHEETS", ""),
    ("1", "README - This sheet"),
    ("2", "Assumptions - Key modelling assumptions"),
    ("3", "Approved_Annual_Budget - Official approved budget (static)"),
    ("4", "Monthly_Budget_Allocation - Monthly allocation profiles"),
    ("5", "Supplementary_Budget - Supplementary budget adjustments"),
    ("6", "Budget_Reallocation - Budget reallocation requests and approvals"),
    ("7", "Expenditure_Transactions - Detailed expenditure transactions"),
    ("8", "Commitments - Outstanding commitments register"),
    ("9", "Budget_Review_Requests - AIE and budget review workflow"),
    ("10","Procurement_Plan - Procurement plan and contract status"),
    ("11","Pending_Bills - Outstanding supplier invoices"),
    ("12","Revenue_AIA - Revenue and Appropriations-in-Aid"),
    ("13","Performance_Indicators - KPI tracking and M&E data"),
    ("14","Audit_Control_Exceptions - Internal audit and control findings"),
    ("15","Data_Quality_Log - Data quality issues and resolutions"),
    ("16","Reconciliation_Control - Data reconciliation checks"),
    ("17","Dashboard_Input - Summary KPIs for dashboard"),
    ("18","Management_Actions - Management action register"),
    ("19","Lookup_Cost_Centres - Cost centre dimension table"),
    ("20","Lookup_Votes - Vote code dimension table"),
    ("21","Lookup_Activities - Activity dimension table"),
    ("22","Lookup_Regions - Region dimension table"),
    ("23","Lookup_Fund_Sources - Fund source dimension table"),
]
for r_idx, (label, value) in enumerate(content, 1):
    c1 = ws_readme.cell(row=r_idx, column=1, value=label)
    c2 = ws_readme.cell(row=r_idx, column=2, value=value)
    if r_idx == 1:
        c1.font = Font(name="Calibri", bold=True, size=14, color=FOREST_GREEN)
        c2.font = Font(name="Calibri", bold=True, size=14, color=MUTED_GOLD)
    elif label and label not in ("","WORKSHEETS"):
        c1.font = Font(name="Calibri", bold=True, size=10, color=CHARCOAL)
        c2.font = Font(name="Calibri", size=10, color=CHARCOAL)
    else:
        c2.font = Font(name="Calibri", size=9, color=CHARCOAL)
    c2.alignment = Alignment(wrap_text=True)
ws_readme.column_dimensions["A"].width = 28
ws_readme.column_dimensions["B"].width = 80
ws_readme.sheet_view.showGridLines = False

# ── 2. Assumptions ──────────────────────────────────────────────────────────
ws_assum = wb.create_sheet("Assumptions")
ws_assum.sheet_view.showGridLines = False
assumptions = [
    ("Parameter","Value","Justification"),
    ("Random Seed","42","Ensures reproducible data generation"),
    ("Financial Year","1 July to 30 June","Kenyan Government financial year"),
    ("Active Financial Year","FY 2025/2026","Current year for analysis"),
    ("Analysis Cut-off Date","31 May 2026","11 months of actual data"),
    ("Total Approved Budget FY2025/26","KSh 1,700,000,000","Realistic parastatal budget scale"),
    ("Total Approved Budget FY2024/25","KSh 1,550,000,000","Prior year benchmark"),
    ("Total Approved Budget FY2023/24","KSh 1,400,000,000","Base year benchmark"),
    ("Number of Cost Centres","32","Spread across 13 directorates and 8 regions"),
    ("Number of Transactions (FY2025/26)","4,500","Generates ~12,100 across 3 FYs"),
    ("Budget Absorption Target (Time-based)","91.7%","11/12 months elapsed"),
    ("Actual YTD Absorption","87.0%","As at 31 May 2026"),
    ("Forecast Method","Blended (30% SL, 25% MA3, 25% Seasonal, 20% Commit-adj)","Multiple methods for robustness"),
    ("Forecast Year-End Expenditure","KSh 1,763,490,330","Slight over-expenditure risk"),
    ("Data Quality Issue Rate","~2-5%","Intentional for demonstration"),
    ("Seasonality","Increased spend Q3/Q4 and June year-end","Reflects public sector spending patterns"),
    ("Currency","Kenya Shillings (KSh)","All amounts in full KSh"),
    ("Tax Rate","0% or 16% VAT","Applied to selected transactions"),
    ("Supplier Data","Fictional IDs and categories only","No real suppliers used"),
]
for r_idx, row in enumerate(assumptions, 1):
    for c_idx, val in enumerate(row, 1):
        cell = ws_assum.cell(row=r_idx, column=c_idx, value=val)
        if r_idx == 1:
            cell.fill = hdr_fill()
            cell.font = hdr_font()
            cell.alignment = center_align()
        else:
            cell.fill = PatternFill("solid", fgColor=LIGHT_GREY if r_idx%2==0 else WHITE)
            cell.font = body_font()
        cell.border = thin_border()
ws_assum.column_dimensions["A"].width = 35
ws_assum.column_dimensions["B"].width = 35
ws_assum.column_dimensions["C"].width = 55

# ── Load all datasets ────────────────────────────────────────────────────────
print("  Loading datasets...")
dfs = {
    "budget":  pd.read_csv(f"{DATA}/Approved_Annual_Budget.csv"),
    "alloc":   pd.read_csv(f"{DATA}/Monthly_Budget_Allocation.csv"),
    "supp":    pd.read_csv(f"{DATA}/Supplementary_Budget.csv"),
    "realloc": pd.read_csv(f"{DATA}/Budget_Reallocation.csv"),
    "txn":     pd.read_csv(f"{DATA}/Expenditure_Transactions.csv"),
    "commit":  pd.read_csv(f"{DATA}/Commitments.csv"),
    "req":     pd.read_csv(f"{DATA}/Budget_Review_Requests.csv"),
    "proc":    pd.read_csv(f"{DATA}/Procurement_Plan.csv"),
    "pb":      pd.read_csv(f"{DATA}/Pending_Bills.csv"),
    "rev":     pd.read_csv(f"{DATA}/Revenue_AIA.csv"),
    "pi":      pd.read_csv(f"{DATA}/Performance_Indicators.csv"),
    "exc":     pd.read_csv(f"{DATA}/Audit_Control_Exceptions.csv"),
    "dq":      pd.read_csv(f"{DATA}/Data_Quality_Log.csv"),
    "cc":      pd.read_csv(f"{DATA}/Dim_Cost_Centre.csv"),
    "votes":   pd.read_csv(f"{DATA}/Dim_Vote.csv"),
    "acts":    pd.read_csv(f"{DATA}/Dim_Activity.csv"),
    "regions": pd.read_csv(f"{DATA}/Dim_Region.csv"),
    "fs":      pd.read_csv(f"{DATA}/Dim_Fund_Source.csv"),
}

print("  Adding data sheets (this may take a moment)...")

# ── 3-7: Core budget and expenditure sheets ──────────────────────────────────
cur_cols_budget = ["ApprovedAnnualAmount"]
add_df_sheet(wb, "Approved_Annual_Budget", dfs["budget"],
             currency_cols=cur_cols_budget, table_name="Tbl_ApprovedBudget")

add_df_sheet(wb, "Monthly_Budget_Allocation", dfs["alloc"],
             currency_cols=["MonthlyAllocationAmount","CumulativeAllocationAmount"],
             table_name="Tbl_MonthlyAllocation")

add_df_sheet(wb, "Supplementary_Budget", dfs["supp"],
             currency_cols=["OriginalApprovedAmount","SupplementaryIncrease",
                            "SupplementaryReduction","NetAdjustment","RevisedApprovedAmount"],
             table_name="Tbl_SupplementaryBudget")

add_df_sheet(wb, "Budget_Reallocation", dfs["realloc"],
             currency_cols=["ReallocationAmount"], table_name="Tbl_Reallocation")

add_df_sheet(wb, "Expenditure_Transactions", dfs["txn"],
             currency_cols=["GrossAmount","TaxAmount","NetAmount"],
             table_name="Tbl_Expenditure")

add_df_sheet(wb, "Commitments", dfs["commit"],
             currency_cols=["CommitmentAmount","AmountInvoiced","AmountPaid","OutstandingCommitment"],
             table_name="Tbl_Commitments")

add_df_sheet(wb, "Budget_Review_Requests", dfs["req"],
             currency_cols=["RequestedAmount","ApprovedAmount","ProcessedAmount","OutstandingAmount"],
             table_name="Tbl_BudgetRequests")

add_df_sheet(wb, "Procurement_Plan", dfs["proc"],
             currency_cols=["EstimatedCost","ApprovedBudget","ActualContractValue",
                            "AmountPaid","OutstandingContractAmount"],
             table_name="Tbl_Procurement")

add_df_sheet(wb, "Pending_Bills", dfs["pb"],
             currency_cols=["InvoiceAmount","AmountVerified","AmountPaid","OutstandingAmount"],
             table_name="Tbl_PendingBills")

add_df_sheet(wb, "Revenue_AIA", dfs["rev"],
             currency_cols=["BudgetedRevenue","ActualRevenue","Variance"],
             table_name="Tbl_Revenue")

add_df_sheet(wb, "Performance_Indicators", dfs["pi"],
             currency_cols=["BudgetAllocated","Expenditure","CostPerOutput"],
             table_name="Tbl_Performance")

add_df_sheet(wb, "Audit_Control_Exceptions", dfs["exc"],
             currency_cols=["AmountAtRisk"], table_name="Tbl_AuditExceptions")

add_df_sheet(wb, "Data_Quality_Log", dfs["dq"], table_name="Tbl_DataQuality")

# ── Reconciliation Control ───────────────────────────────────────────────────
rec_data = pd.read_csv(f"{DATA}/../04_Analysis/Reconciliation_Results.csv") if os.path.exists(f"{DATA}/../04_Analysis/Reconciliation_Results.csv") else pd.DataFrame()
if not rec_data.empty:
    add_df_sheet(wb, "Reconciliation_Control", rec_data, table_name="Tbl_Reconciliation")

# ── Dashboard Input (KPI Summary) ───────────────────────────────────────────
print("  Building Dashboard_Input sheet...")
txn = dfs["txn"]
bud = dfs["budget"]
com = dfs["commit"]
pb  = dfs["pb"]
req = dfs["req"]
rev = dfs["rev"]
pi  = dfs["pi"]
exc = dfs["exc"]

fy = "FY2025/2026"
approved_budget = bud[bud.FinancialYear==fy]["ApprovedAnnualAmount"].sum()
ytd_exp = txn[txn.FinancialYear==fy]["NetAmount"].sum()
total_commits = com[com.FinancialYear==fy]["OutstandingCommitment"].sum()
pending_bills = pb[pb.FinancialYear==fy]["OutstandingAmount"].sum()
pending_reqs  = req[(req.FinancialYear==fy)&(req.Status.str.contains("Pending|Awaiting|Received|Approved but",na=False))]["OutstandingAmount"].sum()
budgeted_rev  = rev[rev.FinancialYear==fy]["BudgetedRevenue"].sum()
actual_rev    = rev[rev.FinancialYear==fy]["ActualRevenue"].sum()
rev_achievement = actual_rev / budgeted_rev * 100 if budgeted_rev > 0 else 0
forecast_ye   = 1_763_490_330
available_bal = approved_budget - ytd_exp - total_commits
util_pct      = ytd_exp / approved_budget * 100
exc_open      = exc[(exc.FinancialYear==fy)&(exc.Status.isin(["Open","In Progress","Escalated"]))]["AmountAtRisk"].sum()

kpis = [
    ("KPI","Value","Unit","Status","Commentary"),
    ("Approved Annual Budget", approved_budget, "KSh", "Reference", "FY2025/2026 original approved budget"),
    ("YTD Actual Expenditure", ytd_exp, "KSh", "Green" if util_pct<=92 else "Amber", f"As at 31 May 2026 ({util_pct:.1f}% utilisation)"),
    ("Budget Utilisation %", util_pct/100, "%", "Green" if util_pct>=80 else "Amber", "YTD vs Approved Budget"),
    ("Total Open Commitments", total_commits, "KSh", "Amber", "Outstanding commitment exposure"),
    ("Available Balance (net of commits)", available_bal, "KSh", "Green" if available_bal>0 else "Red", "After YTD spend and commitments"),
    ("Forecast Year-End Expenditure", forecast_ye, "KSh", "Amber", "Blended forecast - slight over-exp risk"),
    ("Forecast Absorption Rate", forecast_ye/approved_budget, "%", "Amber", "Forecast vs Approved Budget"),
    ("Pending Bills Outstanding", pending_bills, "KSh", "Amber", "Uncleared supplier invoices"),
    ("Budgeted Revenue", budgeted_rev, "KSh", "Reference", "Revenue target FY2025/2026"),
    ("Actual Revenue Collected", actual_rev, "KSh", "Amber" if rev_achievement<90 else "Green", f"Achievement: {rev_achievement:.1f}%"),
    ("Revenue Achievement %", rev_achievement/100, "%", "Amber" if rev_achievement<90 else "Green", "Actual vs Budgeted Revenue"),
    ("Open Audit Exceptions (Amount at Risk)", exc_open, "KSh", "Red" if exc_open>50_000_000 else "Amber", "Unresolved audit and control findings"),
    ("Data Quality Score", 0.962, "%", "Green", "Based on 250 DQ log records, 96.2% clean"),
    ("Pending Approval Requests", pending_reqs, "KSh", "Amber", "Awaiting processing or approval"),
    ("Months Elapsed in FY", 11, "Months", "Reference", "Out of 12 months (Jul 2025 - May 2026)"),
    ("Expected Utilisation (Time-based)", 11/12, "%", "Reference", "91.7% expected at month 11"),
    ("Absorption Gap", (util_pct-91.7)/100, "% pts", "Amber" if abs(util_pct-91.7)>10 else "Green", f"Actual minus expected utilisation"),
]
ws_dash = wb.create_sheet("Dashboard_Input")
ws_dash.sheet_view.showGridLines = False
for r_idx, row in enumerate(kpis, 1):
    for c_idx, val in enumerate(row, 1):
        cell = ws_dash.cell(row=r_idx, column=c_idx, value=val)
        if r_idx == 1:
            cell.fill = hdr_fill()
            cell.font = hdr_font()
            cell.alignment = center_align()
        else:
            status = row[3] if len(row)>3 else ""
            bg = (LIGHT_GREEN if status=="Green" else
                  LIGHT_GOLD if status=="Amber" else
                  "FADBD8" if status=="Red" else WHITE)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = body_font()
        cell.border = thin_border()
        if c_idx == 2 and r_idx > 1:
            v = row[1]
            u = row[2] if len(row)>2 else ""
            if u == "KSh":
                cell.number_format = '#,##0'
            elif u == "%":
                cell.number_format = '0.0%'
ws_dash.column_dimensions["A"].width = 45
ws_dash.column_dimensions["B"].width = 22
ws_dash.column_dimensions["C"].width = 10
ws_dash.column_dimensions["D"].width = 12
ws_dash.column_dimensions["E"].width = 55

# ── Management Actions ───────────────────────────────────────────────────────
actions = [
    ("ActionID","Finding","Risk","RecommendedAction","ResponsibleRole","Priority","TargetDate","Status","ProgressPct","Comments"),
    ("ACT-001","5 cost centres at over-expenditure risk","High","Restrict expenditure and seek reallocation","Director, Finance and Accounts","High","2026-06-15","In Progress",40,"Monthly review initiated"),
    ("ACT-002","KSh 700M open commitments outstanding","High","Review and clear overdue commitments","Director, Supply Chain Management","High","2026-06-20","In Progress",30,"Supplier negotiations ongoing"),
    ("ACT-003","Revenue below target by 12%","Medium","Intensify collection in Q4","Revenue Manager","Medium","2026-06-30","Pending",0,"Revenue drive planned"),
    ("ACT-004","18 cost centres with low absorption","High","Accelerate expenditure on priority activities","Cost Centre Manager","High","2026-06-30","In Progress",20,"Work plans revised"),
    ("ACT-005","Pending bills over 90 days","High","Clear all bills over 90 days priority","Chief Accountant","High","2026-06-10","In Progress",50,"Payment schedule prepared"),
    ("ACT-006","23 cost centres RAG Red on absorption","Critical","Immediate management intervention","Chief Executive Officer","Critical","2026-06-15","In Progress",15,"CEO directive issued"),
    ("ACT-007","Audit exceptions - missing documents","Medium","Obtain and file all missing documents","Head of Internal Audit","Medium","2026-06-30","Pending",0,"Circular issued to all units"),
    ("ACT-008","Procurement delays - 12 contracts delayed","Medium","Issue contract extension or re-procure","Director, Supply Chain Management","Medium","2026-06-30","Pending",0,"Review scheduled"),
    ("ACT-009","Budget requests unprocessed >30 days","Medium","Clear backlog and improve turnaround","Head of Budget","Medium","2026-06-30","In Progress",35,"Processing team reinforced"),
    ("ACT-010","Donor programme underspend","High","Accelerate donor fund utilisation","Project Manager","High","2026-06-30","Pending",0,"Donor notified"),
]
ws_actions = wb.create_sheet("Management_Actions")
for r_idx, row in enumerate(actions, 1):
    for c_idx, val in enumerate(row, 1):
        cell = ws_actions.cell(row=r_idx, column=c_idx, value=val)
        if r_idx == 1:
            cell.fill = hdr_fill(MUTED_GOLD)
            cell.font = Font(name="Calibri", bold=True, size=10, color=CHARCOAL)
            cell.alignment = center_align()
        else:
            priority = row[5] if len(row)>5 else ""
            bg = ("FADBD8" if priority=="Critical" else
                  LIGHT_GOLD if priority=="High" else
                  LIGHT_GREY if priority=="Medium" else WHITE)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = body_font()
        cell.border = thin_border()
ws_actions.freeze_panes = "A2"
for col in ws_actions.columns:
    ws_actions.column_dimensions[get_column_letter(col[0].column)].width = 30

# ── Lookup tables ────────────────────────────────────────────────────────────
add_df_sheet(wb, "Lookup_Cost_Centres", dfs["cc"], table_name="Tbl_CostCentres")
add_df_sheet(wb, "Lookup_Votes", dfs["votes"], table_name="Tbl_Votes")
add_df_sheet(wb, "Lookup_Activities", dfs["acts"], table_name="Tbl_Activities")
add_df_sheet(wb, "Lookup_Regions", dfs["regions"], table_name="Tbl_Regions")
add_df_sheet(wb, "Lookup_Fund_Sources", dfs["fs"], table_name="Tbl_FundSources")

# ── Save ─────────────────────────────────────────────────────────────────────
filepath = f"{OUT}/Parastatal_X_Financial_Monitoring_Data.xlsx"
wb.save(filepath)
size = os.path.getsize(filepath) / (1024*1024)
print(f"\nExcel workbook saved: {filepath}")
print(f"File size: {size:.1f} MB")
print(f"Sheets: {len(wb.sheetnames)}")
for s in wb.sheetnames:
    print(f"  - {s}")
